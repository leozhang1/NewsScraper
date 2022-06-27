import concurrent.futures
import os
import time
from secrets import Secrets

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def getNewsAsDataFrame(category=''):
    url = f'https://newsapi.org/v2/top-headlines?country=us&category={category}&apiKey={Secrets.newsapiKey}'
    newsJson = requests.get(url).json()
    articles = newsJson['articles']
    titles = []
    urls = []
    for i, article in enumerate(articles):
        titles.append(article['title'])
        urls.append(article['url'])
    return pd.DataFrame({'TITLE': titles, 'URL': urls})

def addHyperlinks(filename):
    wb = load_workbook(filename)
    # go thru all sheets and apply hyperlink formula to every row in column B
    for i, sheetname in enumerate(wb.sheetnames):
        ws = wb[sheetname]
        # 1 to n because we're skipping the header
        n = len(wb[sheetname][get_column_letter(1)])
        # make a header for the new column
        ws.cell(row=1, column=3).value = 'Hyperlink'
        # bold the header
        ws.cell(row=1, column=3).font = Font(bold=True)
        # TODO clear out column 1 and 2 but have the user decide that with a boolean

        # go thru every row
        for x in range(2, n+1):
            # print(type(ws.cell(row=x, column=1).value), ws.cell(row=x, column=1).value)
            ws.cell(
                row=x, column=3).value = f"=HYPERLINK({get_column_letter(2)}{x},{get_column_letter(1)}{x})"
    wb.save(filename)

def processToExcelFile() -> str:
    # https://newsapi.org/docs/endpoints/top-headlines
    possibleCategories = ['business', 'entertainment',
                          'general', 'health', 'science', 'sports', 'technology']
    dfs = []
    filePath = f'{os.getcwd()}/files/'
    # didn't use relative path (such as ./files because this path is global on the computer and would maintain correctness of desired file storing locations regardless of who is running this file, whether that's running this script directly or running this script via another python script)
    if not os.path.isdir(filePath):
        os.mkdir(filePath)
    fileName = f'news_{time.strftime("%Y%m%d-%H%M%S")}.xlsx'
    filePathComplete = f'{filePath}{fileName}'
    with concurrent.futures.ThreadPoolExecutor() as executor:
        res = executor.map(getNewsAsDataFrame, possibleCategories)
        # print(list(res))
        dfs.extend(list(res))

    xlw = pd.ExcelWriter(filePathComplete)
    # for category in possibleCategories:
    #     dfs.append(getNewsAsDataFrame(category))
    # print(os.listdir('./CSVs'))
    # merge them into one excel file separated into different sheets
    for i, df in enumerate(dfs):
        df.to_excel(xlw, sheet_name=possibleCategories[i], index=False)
    xlw.close()
    return filePath, fileName

def main(shouldDeleteFile=False):
    lastDate = ''
    if not os.path.isfile(f'{os.getcwd()}/time_stamp.txt'):
        print('creating file')
        with open(f'{os.getcwd()}/time_stamp.txt', 'w') as f:
            f.write('')
    with open(f'{os.getcwd()}/time_stamp.txt', 'r') as f:
        lastDate = f.read()
        if lastDate == time.strftime("%Y-%m-%d"):
            print('already ran this')
            return
    with open(f'{os.getcwd()}/time_stamp.txt', 'w') as f:
        f.write(time.strftime("%Y-%m-%d"))


    filePath, fileName = processToExcelFile()
    addHyperlinks(f'{filePath}{fileName}')

    if shouldDeleteFile and os.path.isfile(filePath+fileName):
        os.remove(filePath+fileName)


if __name__ == '__main__':
    os.chdir(os.path.dirname(__file__))
    # print(os.getcwd())
    main(shouldDeleteFile=False)


# region write to csv
# def writeNewsToFile(titles=[], urls=[]):
    # region csv code
    # with open(f'./CSVs/{category}_category_news_{time.strftime("%Y%m%d-%H%M%S")}.csv', 'a', encoding='UTF8', newline='') as f:
    #     w = csv.writer(f)
    #     w.writerow(['TITLE','.', 'URL'])
    #     for title,url in newsLst:
    #         w.writerow((title,'.',url))
    # endregion
    # convert to pandas df
#     pass
# endregion

# region tkinter code
# ? window config
# root = tk.Tk()
# root.geometry('')
# root.title('News')
# scrollFrame = Frame(root)
# scrollFrame.pack(fill=BOTH, expand=1)
# canvas = Canvas(scrollFrame)
# canvas.pack(side=LEFT, fill=BOTH,expand=1)
# scrollbar = tk.Scrollbar(scrollFrame, orient=VERTICAL, command=canvas.yview)
# scrollbar.pack(side=RIGHT, fill=Y)
# #configure canvas for scrollbar
# canvas.configure(yscrollcommand=scrollbar.set)
# canvas.bind('<Configure>', lambda event: canvas.configure(scrollregion=canvas.bbox('all')))
# secondFrame = Frame(canvas,bg='gray')
# secondFrame.pack(padx=500,pady=500)
# canvas.create_window((0,0), window=secondFrame, anchor='nw')
# button = tk.Button(secondFrame, font=24, text='Reload', command=getNews,)
# button.pack()
# label = tk.Label(secondFrame, font=24, justify="left",anchor=NW, background='black', foreground='white')
# # pack on to the window
# label.pack(side=BOTTOM)
# label.config(text=getNews())
# thirdFrame = Frame(canvas,bg='gray')
# canvas.create_window((150,100), window=thirdFrame)
# root.mainloop()
# endregion
